import requests
import bs4
import os
import re
import pprint
import json
import hashlib
import openpyxl
import datetime
import time

from tqdm import tqdm
from selenium import webdriver

from CoinDict import *
from ChanScapes import *
from mongo_db.tickerTable import *


# Generate list of coins
generateCurrenciesList()    
# Load json file for coin list 
with open('data.json') as json_file:
    CD = json.load(json_file)

def ThreadIDGet():
    try:
        os.remove('tids.txt')
    except:
        print("Did NOT delete tids.txt")
    url = 'http://boards.4chan.org/biz/catalog'
    driver = webdriver.Chrome()
    driver.get(url)
    res = driver.page_source
    soup = bs4.BeautifulSoup(res, 'lxml')
    body = soup.find('body')
    content = body.find(id="content")
    thread = content.find(id="threads")
    ThR = re.compile(r'(thread-)(\d\d\d\d\d\d\d\d)')

    threadlist = ThR.findall(str(thread))
    ThreadIDs = open("tids.txt", 'a')
    for i in range(1, len(threadlist)):
        ThreadIDs.write(threadlist[i][1] + '\n')
    print('IDs Obtained')
    driver.close()
    ThreadIDs.close()

def TextGet():
    try:
        os.remove('text.txt')
    except:
        print("Did NOT delete text.txt")
    tids = open('tids.txt','r')
    tlist = tids.readlines()

    print("Scanning threads")
    tickerDb = MongoDB_Biz_Ticker_Mentions()
    for i in tqdm(range(0, len(tlist)-1)):
        url = 'http://boards.4chan.org/biz/thread/' + tlist[i]

        # threadJson = fullThreadScrape(tlist[i][:-2],url)
        threadJson = tickerOnlyScrape(tlist[i],url,tickerDb)

    print('Scrape Complete')

def Count(row):
    file = open('text.txt', 'r')
    postlist = file.readlines()
    checklist = [row['aka'][0],row['aka'][0].lower(),(row['name']),(row['name'].lower())]
    Count = 0
    print(checklist)
    for i in range(0,len(postlist)):
        for x in checklist:
            if x in postlist[i].split():
                Count += 1
                break
            else:
                continue
    file.close()
    return Count

def NewBook(Name):
    wb = openpyxl.Workbook()
    print(type(wb))
    inp = wb.active
    print(type(inp))
    inp['B1']='TIME'
    inp['A2']='Coin'
    inp['B2'] = 'Metric'
    inp.freeze_panes = 'C1'

    x = 0
    for row in CD:
        try:
            inp['A{}'.format(str(x + 3))] = row['name']
            inp['B{}'.format(str(x + 3))] = 'BVol'
            inp['B{}'.format(str(x + 4))] = 'Price'
            inp['B{}'.format(str(x + 5))] = 'Posts'
            x += 5
        except:
            continue
    #for i in range(0,90):


    wb.save(Name + '.xlsx')

def Update():
    emptychk = 3
    x = 3
    wb = openpyxl.load_workbook('tester.xlsx')
    inp = wb.active
    refC = inp.cell(row = 3,column = emptychk)
    while refC.value != None:
        emptychk += 1
        refC = inp.cell(row=3, column=emptychk)
    inp.cell(row = 1, column = emptychk).value = datetime.datetime.now()
    #Update Volume
    print('UPDATING')
    for row in CD:
        try:
            data = get24HrStats(row["aka"][0].upper())
            data = data["data"]
            volume = data['vol']
            price = data['last']
            VCell = inp.cell(row=x, column=emptychk)
            PCell = inp.cell(row=x + 1, column=emptychk)
            CCell = inp.cell(row=x + 2, column=emptychk)
            VCell.value = volume
            PCell.value = price
            CCell.value = Count(row)
            #print(str(row['aka'][0]))
            #print(str(CCell.value) + ' writing to: ' + str(emptychk) + ' , ' + str(x + 2))
            x += 5
        except:
            print(str(row['aka'][0]) + ' Not Found')
            continue
        
    print('UPDATED')
    wb.save('tester.xlsx')

NewBook('tester')
while True:
    
    while True:
        ThreadIDGet()
        TextGet()
    Update()


